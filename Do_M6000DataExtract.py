# -*- coding: utf-8 -*-
# This program is developed with python 3.x
# The following two types of input are avaliable:
#     1. Directory path of M6000 .dat files
#     2. the List of M6000 .dat files

# from PyQt5 import QtCore, QtGui, QtWidgets

import datetime
import re
import os

# workbook
from  openpyxl.workbook  import  Workbook  
# ExcelWriter
from  openpyxl.writer.excel  import  ExcelWriter  
# convert number to column alpha-belta
# from  openpyxl.cell  import  get_column_letter

#class extractM6000Cfg(QtCore.QThread):
class extractM6000Cfg():
    """ class extractM6000Cfg """

    # Patterns
    ptnInterface        = re.compile(r'\s*interface\s*(\S*)')
    ptnInterfaceEnd     = re.compile(r'\s*\$')

    ptnVlanID           = re.compile(r'\s*qinq internal-vlanid\s*(\S*)\s*external-vlanid\s*(\S*)')
    ptnVlanRange        = re.compile(r'\s*qinq range internal-vlan-range\s*(\S*)\s*external-vlan-range\s*(\S*)')
    ptnDescription      = re.compile(r'\s*description\s*(\S*)')
    ptnForwarding       = re.compile(r'\s*ip vrf forwarding\s*(\S*)')
    ptnIPAddress        = re.compile(r'\s*ip address\s*(\S*)\s*(\S*)')

    

    ptnDict = {  'description'  : ptnDescription, \
                 'forwarding'   : ptnForwarding, \
                 'IPAddress'    : ptnIPAddress, \
                 'vlanID'       : ptnVlanID, \
                 'vlanRange'    : ptnVlanRange }

    # signal
    # sigProcessFiles = QtCore.pyqtSignal(str)
    # sigRecord = QtCore.pyqtSignal(str)

#    sigRecordClear = QtCore.pyqtSignal()
#    sigLineProcessed = QtCore.pyqtSignal(int)
#    sigTotalLineNum = QtCore.pyqtSignal(int)
#    sigStartTimer = QtCore.pyqtSignal()
#    sigStopTimer = QtCore.pyqtSignal()

    def __init__(self):
        super(extractM6000Cfg, self).__init__()
        self.path = None

    def loadPara(self, path, outFile = None):
        self.path = path
        if outFile is None:
            self.outFile = self.genOutFilename()
        else:
            self.outFile = outFile
        print(self.outFile)
        self.workbook = Workbook()
        # self.excelWriter = ExcelWriter(self.workbook)
        # self.worksheet = self.workbook.create_sheet(title="M6000")
        self.worksheet = self.workbook.active
        self.worksheet.title = 'M6000'
        self.worksheet.append(['filename', 'interface', 'description', 'ip vrf forwarding', 'ip', 'mask', \
                'internal-vlanid', 'external-vlanid', 'internal-vlan-range', 'external-vlan-range'])

    def genOutFilename(self):
        t = datetime.datetime.now()
        s = datetime.datetime.strftime(t, '%Y%m%d_%H%M%S')
        filename = 'M6000_' + s + '.xlsx'
        if type(self.path) is str: # input is a directory string
            path = self.path
        elif type(self.path) is list: # input is a file list
            path = os.path.split(self.path[0])[0]
        return os.path.join(path, filename)


    def run(self):
        if self.path is None:
            return
        if type(self.path) is str: # input is a directory string
            self.directoryExtract(self.path)
        elif type(self.path) is list: # input is a file list
            for fn in self.path:
                self.fileExtract(fn)
            #if os.path.exists(self.path):
        # save the output file
        self.saveOutFile()


    def directoryExtract(self, dn):
        if os.path.isdir(dn):
            for parent, dirnames, filenames in os.walk(dn):
                for filename in filenames:
                    fn = os.path.join(parent, filename)
                    self.fileExtract(fn)


    def fileExtract(self, fn):
        # check if file is *.dat
        if os.path.splitext(fn)[-1] != '.dat':
            return

        fnBase = os.path.split(os.path.splitext(fn)[0])[-1]
        print(fn)
        print(fnBase)
        # start analyse
        isInterfaceStart = False
        interfaces = {}
        with open(fn, 'r') as fp:
            for line in fp:
                #line = line.strip()
                if isInterfaceStart:
                    if self.ptnInterfaceEnd.match(line):
                        isInterfaceStart = False
                    else:
                        # match interface param
                        for title, ptn in self.ptnDict.items():
                            result = ptn.match(line)
                            if result:
                                if title == 'vlanID' or title == 'vlanRange' or title == 'IPAddress':
                                    interfaces[ifname][title] = [result.group(1), result.group(2)]
                                else:
                                    # debug end
                                    interfaces[ifname][title] = result.group(1)
                                break
                else:
                    #match interface internal parameters
                    result = self.ptnInterface.match(line)
                    if result:
                        ifname = result.group(1)
                        if interfaces.get(ifname):
                            pass
                        else:
                            interfaces[ifname] = {}
                        isInterfaceStart = True

        # file analyse finished, save the valid records
        for ifname, params in interfaces.items():
            if params.get('IPAddress'):
                self.saveItem(fnBase, ifname, params)

    def saveOutFile(self):
        self.workbook.save(self.outFile)

    def saveItem(self, fnBase, ifname, params):
        self.writeItemToExcel(fnBase, ifname, params)

    def checkAndFormat(self, item): # TODO
        # interface name must contains '.'
        if len(item.get('interface', '')) <= 1:
            return False
        # format interface description
        interfaceDescription = item['interface'].get('description') 
        if interfaceDescription:
            try:
                result = re.match(r'\w\d{7}', interfaceDescription.split()[0])
                if result:
                    item['interface']['description header'] = result.group(0)
            except Exception as inst:
                print(inst)
                pass
        # format secondary ip address
        ipList = item['interface'].get('secondary')
        if ipList is not None:
            item['interface']['secondary'] = ' '.join(ipList)
        return True

    def writeItemToExcel(self, fnBase, ifname, params):
        vlanidLst = params.get('vlanID', ['',''])
        vlanRangeLst = params.get('vlanRange', ['',''])
        self.worksheet.append( [\
                                fnBase, \
                                ifname, \
                                params.get('description', ''), \
                                params.get('forwarding', ''), \
                                params['IPAddress'][0], \
                                params['IPAddress'][1], \
                                vlanidLst[0], \
                                vlanidLst[1], \
                                vlanRangeLst[0], \
                                vlanRangeLst[1] \
                                ])
        #, 'ip-address sub')
        pass

    def fRecord(self, s):
        #self.sigRecord.emit(s)
        print(s)

    def fNowPastTimeStr(self):
        delta = datetime.datetime.now() - self.tAnalyseStartTime
        return '%d.%d seconds' % (delta.seconds, delta.microseconds)


    def fGetFileLines(self, fn):
        i = -1
        with open(fn) as f:
            for i, l in enumerate(f):
                pass
        return i + 1

    def fSummary(self):
        s = []
        s.append( 'Summary:' )
        self.fRecord('\r\n'.join(s))

    def fStartTimer(self):
        self.sigStartTimer.emit()

    def fTimerEvent(self):
        self.sigLineProcessed.emit(self.lineProcess)

    def fStopTimer(self):
        self.sigStopTimer.emit()
        self.sigLineProcessed.emit(self.lineProcess)


if __name__ == "__main__":
    fn = input('Input file:')
    app = extractM6000Cfg()
    #app.sigRecord.connect(print)
    if os.path.isdir(fn):
        inParam = fn
    else:
        inParam = [fn]
    app.loadPara(inParam)
    app.run()
    input('')
