# -*- coding: utf-8 -*-
# This program is developed with python 3.x
# The following two types of input are avaliable:
#     1. Directory path of VPN .cfg files
#     2. the List of VPN .cfg files

#from PyQt5 import QtCore, QtGui, QtWidgets

import datetime
import re
import os

# workbook
from  openpyxl.workbook  import  Workbook  
# ExcelWriter
from  openpyxl.writer.excel  import  ExcelWriter  
# convert number to column alpha-belta
# from  openpyxl.cell  import  get_column_letter

#class extractVPNCfg(QtCore.QThread):
class extractVPNCfg():
    """ class extractVPNCfg """

    ptnInterface = re.compile(r'interface\s*(\S*)')
    ptnShutdown = re.compile(r'shutdown')
    ptnDescription = re.compile(r'description\s*(\S*)')
    ptnTermination = re.compile(r'q termination\s*(\S.*)')
    ptnVPNInstance = re.compile(r'vpn-instance\s*(\S*)')
    ptnIPAddress = re.compile(r'ip address\s*(\S.*)')
    ptnQosInbound = re.compile(r'qos-profile\s*(\d\S*)\s*inbound')
    ptnQosOutbound = re.compile(r'qos-profile\s*(\d\S*)\s*outbound')

    paramPtnDict = { 'description' : ptnDescription, \
                     'shutdown' : ptnShutdown, \
                     'termination' : ptnTermination, \
                     'vpn-instance' : ptnVPNInstance, \
                     'qos inbound' : ptnQosInbound, \
                     'qos outbound' : ptnQosOutbound, \
                     'ip address' : ptnIPAddress }

    # signal
    #sigProcessFiles = QtCore.pyqtSignal(str)
    # sigRecord = QtCore.pyqtSignal(str)
#    sigRecordClear = QtCore.pyqtSignal()
#    sigLineProcessed = QtCore.pyqtSignal(int)
#    sigTotalLineNum = QtCore.pyqtSignal(int)
#    sigStartTimer = QtCore.pyqtSignal()
#    sigStopTimer = QtCore.pyqtSignal()

    def __init__(self):
        super(extractVPNCfg, self).__init__()
        self.path = None

    def loadPara(self, path, outFile = None):
        self.path = path
        if outFile is None:
            self.outFile = self.genOutFilename()
        else:
            self.outFile = outFile
        print(self.outFile)
        self.workbook = Workbook()
        # self.excelWriter = ExcelWriter(self.workbook)
        # self.worksheet = self.workbook.create_sheet(title="VPN")
        self.worksheet = self.workbook.active
        self.worksheet.title = 'VPN'
#        self.worksheet.append(['ip', 'interface', 'description', 'termination', 'vpn-instance', \
#                'ip address', 'netmask', 'ip-address sub', 'netmask sub'])
        self.worksheet.append(['ip', 'interface-1', 'interface-2', 'interface-3', 'description', 'is shutdown', 'pe-vid', 'ce-vid', 'vpn-instance', \
                'ip address', 'netmask', 'ip-address sub', 'netmask sub', 'qos inbound', 'qos outbound'])

    def genOutFilename(self):
        t = datetime.datetime.now()
        s = datetime.datetime.strftime(t, '%Y%m%d_%H%M%S')
        filename = 'VPN_' + s + '.xlsx'
        if type(self.path) is str: # input is a directory string
            path = self.path
        elif type(self.path) is list: # input is a file list
            path = os.path.split(self.path[0])[0]
        return os.path.join(path, filename)


    def run(self):
        if self.path is None:
            return
        if type(self.path) is str: # input is a directory string
            self.directoryExact(self.path)
        elif type(self.path) is list: # input is a file list
            for fn in self.path:
                self.fileExact(fn)
            #if os.path.exists(self.path):
        # save the output file
        self.saveOutFile()


    def directoryExact(self, dn):
        if os.path.isdir(dn):
            for parent, dirnames, filenames in os.walk(dn):
                for filename in filenames:
                    fn = os.path.join(parent, filename)
                    self.fileExact(fn)


    def fileExact(self, fn):
        # check if file is *.cfg
        if os.path.splitext(fn)[-1] != '.cfg':
            return

        print('dealing with %s' % fn)
        fnBase = os.path.split(os.path.splitext(fn)[0])[-1]
        if fnBase == 'vrpcfg':
            ip = os.path.splitext(fn)[0].split('\\')[-2]
        else:
            ip = fnBase
        # print(fnBase)
        # start analyse
        isInterfaceStart = False
        item = {}
        with open(fn, 'r') as fp:
            for line in fp:
                #line = line.strip()
                if isInterfaceStart:
                    # reach the end of the instance
                    if line[0] == '#': #reach the end
                        item['ip'] = ip
                        self.saveItem(item)
                        #reset item
                        isInterfaceStart = False
                        item = {}
                        keyItemNum = 0
                    else:
                        #match interface internal parameters
                        for title, ptn in self.paramPtnDict.items():
                            result = ptn.search(line)
                            if result:
                                if title == 'ip address':
                                    if item.get(title) is None:
                                        item[title] = [result.group(1)]
                                    else:
                                        item[title].append(result.group(1))
                                elif title == 'shutdown':
                                    if 'undo' in line:
                                        item[title] = 'N'
                                    else:
                                        item[title] = 'Y'
                                else:
                                    item[title] = result.group(1)
                else: # interface not started, find the start
                    result = self.ptnInterface.match(line) 
                    if result:
                        item['interface'] = result.group(1)
                        isInterfaceStart = True

    def saveOutFile(self):
        self.workbook.save(self.outFile)

    def saveItem(self, item):
        if self.checkAndFormat(item):
            self.writeItemToExcel(item)
        else:
            pass
            # print(item)

    def checkAndFormat(self, item):
        # interface name must contains '.'
        if '.' not in item.get('interface'):
            return False
        # interface example: 'GigabitEthernet1/1/6.47'
        # devide into: 'GigabitEthernet' '1/1/6' '47'
        ifList = item.get('interface').rsplit('.', 1)
        if len(ifList) != 2:
            return false
        matchResult = re.match(r'(\D+)(.*)', ifList[0])
        item['interface-1'] = matchResult.group(1)
        item['interface-2'] = matchResult.group(2)
        item['interface-3'] = ifList[1]
        # one of termination, vpn-instance, ip address MUST be set
        if item.get('termination') is None and item.get('vpn-instance') is None and item.get('ip address') is None:
            return False
        # format termination
        termination = item.get('termination')
        if termination:
            vidList = termination.split()
            if vidList[0] == 'vid':
                item['termination'] = 'pe-vid -1 ce-vid %s' % vidList[1]
        # convert termination to pe-vid & ce-vid
        termination = item.get('termination')
        if termination:
            vidList = termination.split()
            item['pe-vid'] = vidList[1]
            item['ce-vid'] = vidList[3]

        # format ip address
        ipList = item.get('ip address')
        if ipList is not None:
            ip_mask = ipList[0].split()
            item['ip address'] = ip_mask[0]
            item['netmask'] = ip_mask[1]
            if len(ipList) > 1:
                ip_mask = ipList[1].split()
                item['ip address sub'] = ip_mask[0]
                item['netmask sub'] = ip_mask[1]
        return True

    def writeItemToExcel(self, item):
        self.worksheet.append( [\
                                item.get('ip'), \
                                item.get('interface-1'), \
                                item.get('interface-2'), \
                                item.get('interface-3'), \
                                item.get('description', ''), \
                                item.get('shutdown', ''), \
                                #item.get('termination', ''), \
                                item.get('pe-vid', ''), \
                                item.get('ce-vid', ''), \
                                item.get('vpn-instance', ''), \
                                item.get('ip address', ''), \
                                item.get('netmask', ''), \
                                item.get('ip address sub', ''), \
                                item.get('netmask sub', ''), \
                                item.get('qos inbound', ''), \
                                item.get('qos outbound', ''), \
                                ] \
                                )
        #, 'ip-address sub')
        pass


    def fRecord(self, s):
        #self.sigRecord.emit(s)
        print(s)

    def fNowPastTimeStr(self):
        delta = datetime.datetime.now() - self.tAnalyseStartTime
        return '%d.%d seconds' % (delta.seconds, delta.microseconds)


    def fGetFileLines(self, fn):
        i = -1
        with open(fn) as f:
            for i, l in enumerate(f):
                pass
        return i + 1

    def fSummary(self):
        s = []
        s.append( 'Summary:' )
        self.fRecord('\r\n'.join(s))

    def fStartTimer(self):
        self.sigStartTimer.emit()

    def fTimerEvent(self):
        self.sigLineProcessed.emit(self.lineProcess)

    def fStopTimer(self):
        self.sigStopTimer.emit()
        self.sigLineProcessed.emit(self.lineProcess)


if __name__ == "__main__":
    fn = input('Input file:')
    app = extractVPNCfg()
    # app.sigRecord.connect(print)
    if os.path.isdir(fn):
        inParam = fn
    else:
        inParam = [fn]
    app.loadPara(inParam)
    app.run()
    input('')
