# -*- coding: utf-8 -*-
# This program is developed with python 3.x
# The following two types of input are avaliable:
#     1. Directory path of VPRN .txt files
#     2. the List of VPRN .txt files

from PyQt5 import QtCore, QtGui, QtWidgets

import datetime
import re
import os

# workbook
from  openpyxl.workbook  import  Workbook  
# ExcelWriter
from  openpyxl.writer.excel  import  ExcelWriter  
# convert number to column alpha-belta
# from  openpyxl.cell  import  get_column_letter

class extractVPRNCfg(QtCore.QThread):
    """ class extractVPRNCfg """

    ptnVPRN             = re.compile(r'\s{8}vprn\s*(\S*)')
    ptnVPRNDescription  = re.compile(r'\s{12}description\s*"(.*)"')
    ptnVPRNAutonomous   = re.compile(r'\s{12}autonomous-system\s*(\S*)')
    ptnVPRNRouteDist    = re.compile(r'\s{12}route-distinguisher\s*(\S*)')
    ptnInterface        = re.compile(r'\s{12}interface\s*"(.*)"')
    ptnIFDescription    = re.compile(r'\s{16}description\s*"(.*)"')
    ptnIFAddress        = re.compile(r'\s{16}address\s*(\S*)')
    ptnIFSecondary      = re.compile(r'\s{16}secondary\s*(\S*)')
    ptnIFSap            = re.compile(r'\s{16}sap\s*(\S*)')
    ptnIFIngress        = re.compile(r'\s{20}ingress')
    ptnIFEgress         = re.compile(r'\s{20}egress')
    ptnIFQos            = re.compile(r'\s{24}qos\s*(\S*)')
    ptnIFIngressEnd     = re.compile(r'\s{20}exit')
    ptnIFEgressEnd      = re.compile(r'\s{20}exit')
    ptnInterfaceEnd     = re.compile(r'\s{12}exit')
    ptnVPRNEnd          = re.compile(r'\s{8}exit')
    #ptnVPRNInstance      = re.compile(r'vpn-instance\s*(\S*)')
    #ptnIPAddress        = re.compile(r'ip address\s*(\S.*)')

    vprnPtnDict = { 'description' : ptnVPRNDescription, \
                     'route-distinguisher' : ptnVPRNRouteDist, \
                     'autonomous-system' : ptnVPRNAutonomous, \
                     'interface' : ptnInterface }
    ifPtnDict = { 'description' : ptnIFDescription, \
                    'address' : ptnIFAddress, \
                    'secondary' : ptnIFSecondary, \
                    'sap' : ptnIFSap, \
                    'qos' : ptnIFQos, \
                    'ingress' : ptnIFIngress, \
                    'egress' : ptnIFEgress }

    # signal
    # sigProcessFiles = QtCore.pyqtSignal(str)
    sigRecord = QtCore.pyqtSignal(str)

#    sigRecordClear = QtCore.pyqtSignal()
#    sigLineProcessed = QtCore.pyqtSignal(int)
#    sigTotalLineNum = QtCore.pyqtSignal(int)
#    sigStartTimer = QtCore.pyqtSignal()
#    sigStopTimer = QtCore.pyqtSignal()

    def __init__(self):
        super(extractVPRNCfg, self).__init__()
        self.path = None

    def loadPara(self, path, outFile = None):
        self.path = path
        if outFile is None:
            self.outFile = self.genOutFilename()
        else:
            self.outFile = outFile
        print(self.outFile)
        self.workbook = Workbook()
        # self.excelWriter = ExcelWriter(self.workbook)
        # self.worksheet = self.workbook.create_sheet(title="VPRN")
        self.worksheet = self.workbook.active
        self.worksheet.title = 'VPRN'
        self.worksheet.append(['filename', 'vprn', 'description', 'autonomous-system', 'route-distinguisher', \
                'interface:name', 'interface:description', 'interface:description header', 'interface:address', \
                'interface:secondary address', 'interface:sap', 'interface:ingress qos', 'interface:egress qos'])

    def genOutFilename(self):
        t = datetime.datetime.now()
        s = datetime.datetime.strftime(t, '%Y%m%d_%H%M%S')
        filename = 'VPRN_' + s + '.xlsx'
        if type(self.path) is str: # input is a directory string
            path = self.path
        elif type(self.path) is list: # input is a file list
            path = os.path.split(self.path[0])[0]
        return os.path.join(path, filename)


    def run(self):
        if self.path is None:
            return
        if type(self.path) is str: # input is a directory string
            self.directoryExtract(self.path)
        elif type(self.path) is list: # input is a file list
            for fn in self.path:
                self.fileExtract(fn)
            #if os.path.exists(self.path):
        # save the output file
        self.saveOutFile()


    def directoryExtract(self, dn):
        if os.path.isdir(dn):
            for parent, dirnames, filenames in os.walk(dn):
                for filename in filenames:
                    fn = os.path.join(parent, filename)
                    self.fileExtract(fn)


    def fileExtract(self, fn):
        # check if file is *.txt
        if os.path.splitext(fn)[-1] != '.txt':
            return

        fnBase = os.path.split(os.path.splitext(fn)[0])[-1]
        print(fn)
        print(fnBase)
        # start analyse
        isVPRNStart = False
        isInterfaceStart = False # a VPRN may contain several Interfaces.
        item = {}
        with open(fn, 'r') as fp:
            for line in fp:
                #line = line.strip()
                if isVPRNStart:
                    # reach the end of the instance
                    if self.ptnVPRNEnd.match(line):
                        isVPRNStart = False
                        item = {}
                        item['filename'] = fnBase
                    else:
                        if isInterfaceStart:
                            if self.ptnInterfaceEnd.match(line):
                                self.saveItem(item)
                                isInterfaceStart = False
                            else:
                                # match interface param
                                for title, ptn in self.ifPtnDict.items():
                                        result = ptn.match(line)
                                        if result:
                                            if title == 'secondary':
                                                if item['interface'].get(title):
                                                    pass
                                                else:
                                                    item['interface'][title] = []
                                                item['interface'][title].append(result.group(1))
                                            else:
                                                if title == 'ingress' or title == 'egress':
                                                    xgress = title
                                                    break
                                                elif title == 'qos':
                                                    title = xgress
                                                # debug start
                                                if item['interface'].get(title):
                                                    print(title, 'exists')
                                                # debug end
                                                item['interface'][title] = result.group(1)
                                            break
                        else:
                            #match interface internal parameters
                            for title, ptn in self.vprnPtnDict.items():
                                result = ptn.match(line)
                                if result:
                                    if title == 'interface':
                                        item[title] = {'name': result.group(1)}
                                        isInterfaceStart = True
                                    else:
                                        item[title] = result.group(1)
                                    break

                else: # interface not started, find the start
                    result = self.ptnVPRN.match(line) 
                    if result:
                        item['vprn'] = result.group(1)
                        isVPRNStart = True

    def saveOutFile(self):
        self.workbook.save(self.outFile)

    def saveItem(self, item):
        if self.checkAndFormat(item):
            self.writeItemToExcel(item)
        else:
            pass

    def checkAndFormat(self, item): # TODO
        # interface name must contains '.'
        if len(item.get('interface', '')) <= 1:
            return False
        # format interface description
        interfaceDescription = item['interface'].get('description') 
        if interfaceDescription:
            try:
                result = re.match(r'\w\d{7}', interfaceDescription.split()[0])
                if result:
                    item['interface']['description header'] = result.group(0)
            except Exception as inst:
                print(inst)
                pass
        # format secondary ip address
        ipList = item['interface'].get('secondary')
        if ipList is not None:
            item['interface']['secondary'] = ' '.join(ipList)
        return True

    def writeItemToExcel(self, item):
        self.worksheet.append( [\
                                item.get('filename'), \
                                item.get('vprn'), \
                                item.get('description', ''), \
                                item.get('autonomous-system', ''), \
                                item.get('route-distinguisher', ''), \
                                item['interface'].get('name', ''), \
                                item['interface'].get('description', ''), \
                                item['interface'].get('description header', ''), \
                                item['interface'].get('address', ''), \
                                item['interface'].get('secondary', ''), \
                                item['interface'].get('sap', ''), \
                                item['interface'].get('ingress', ''), \
                                item['interface'].get('egress', '')]
                                )
        #, 'ip-address sub')
        pass

    def fRecord(self, s):
        self.sigRecord.emit(s)

    def fNowPastTimeStr(self):
        delta = datetime.datetime.now() - self.tAnalyseStartTime
        return '%d.%d seconds' % (delta.seconds, delta.microseconds)


    def fGetFileLines(self, fn):
        i = -1
        with open(fn) as f:
            for i, l in enumerate(f):
                pass
        return i + 1

    def fSummary(self):
        s = []
        s.append( 'Summary:' )
        self.fRecord('\r\n'.join(s))

    def fStartTimer(self):
        self.sigStartTimer.emit()

    def fTimerEvent(self):
        self.sigLineProcessed.emit(self.lineProcess)

    def fStopTimer(self):
        self.sigStopTimer.emit()
        self.sigLineProcessed.emit(self.lineProcess)


if __name__ == "__main__":
    fn = input('Input file:')
    app = extractVPRNCfg()
    app.sigRecord.connect(print)
    if os.path.isdir(fn):gg
        inParam = fn
    else:
        inParam = [fn]
    app.loadPara(inParam)
    app.start()
    input('')
